import re
from collections import Counter
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook


st.set_page_config(page_title="通航驾驶员能力评估分析", layout="wide")


TOTAL_MAX_SCORE = 125
PART_MAX_SCORES = {
    "知识部分": 40,
    "技能部分": 60,
    "态度部分": 25,
}
PART_ORDER = list(PART_MAX_SCORES.keys())
DEFAULT_EXCEL_PATH = Path(__file__).with_name("通航检查情况整理20260626.xlsx")
PERSON_REQUIRED_COLUMNS = ["人员ID", "姓名", "工作单位"]
PERSON_BASE_COLUMNS = [
    "人员ID",
    "序号",
    "姓名",
    "驾驶员执照编号",
    "工作单位",
    "运行基地",
    "知识部分总分",
    "技能部分总分",
    "态度部分总分",
    "总成绩",
    "原评估结果",
    "规则评估结果",
    "最终评估结果",
    "最终评估结果显示",
    "评估结果一致性",
    "结果差异提醒",
    "安全风险提示",
    "人工安全风险标记",
    "检查员评语",
]
DETAIL_BASE_COLUMNS = [
    "人员ID",
    "姓名",
    "工作单位",
    "运行基地",
    "来源文件",
    "部分",
    "评分项目",
    "题目",
    "得分",
    "满分",
    "得分率",
    "列顺序",
]
RESULT_ORDER = ["优秀", "通过", "临界", "不合格", "未评定"]
RESULT_LABELS = {
    "优秀": "优秀（≥100分）",
    "通过": "通过（≥75分）",
    "临界": "临界（≥62分）",
    "不合格": "不合格（<62分）",
    "未评定": "未评定",
}
RESULT_LABEL_TO_STATUS = {label: status for status, label in RESULT_LABELS.items()}
RESULT_LABEL_ORDER = [RESULT_LABELS[status] for status in RESULT_ORDER]
RESULT_COLOR_MAP = {
    RESULT_LABELS["优秀"]: "#2563EB",
    RESULT_LABELS["通过"]: "#16A34A",
    RESULT_LABELS["临界"]: "#F59E0B",
    RESULT_LABELS["不合格"]: "#DC2626",
    RESULT_LABELS["未评定"]: "#6B7280",
}

CHART_FONT_COLOR = "#111827"
CHART_BASE_FONT_SIZE = 15
CHART_TITLE_FONT_SIZE = 22


def apply_readable_chart_fonts(fig):
    if fig is None:
        return fig
    fig.update_layout(
        font=dict(color=CHART_FONT_COLOR, size=CHART_BASE_FONT_SIZE),
        title_font=dict(color=CHART_FONT_COLOR, size=CHART_TITLE_FONT_SIZE),
        legend=dict(font=dict(color=CHART_FONT_COLOR, size=CHART_BASE_FONT_SIZE)),
        margin=dict(l=32, r=32, t=72, b=40),
    )
    fig.update_xaxes(tickfont=dict(color=CHART_FONT_COLOR, size=CHART_BASE_FONT_SIZE))
    fig.update_yaxes(tickfont=dict(color=CHART_FONT_COLOR, size=CHART_BASE_FONT_SIZE))
    return fig


_original_plotly_chart = st.plotly_chart


def readable_plotly_chart(fig, *args, **kwargs):
    return _original_plotly_chart(apply_readable_chart_fonts(fig), *args, **kwargs)


st.plotly_chart = readable_plotly_chart


def compact_text(value):
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_key(value):
    return re.sub(r"\s+", "", compact_text(value))


def to_number(value):
    if value is None:
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    text = compact_text(value)
    if not text or "空白" in text:
        return np.nan
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else np.nan


def figure_height(rows, minimum=420, per_row=28, maximum=1600):
    return int(min(max(minimum, rows * per_row + 180), maximum))


def parse_max_score(label, default=np.nan):
    text = compact_text(label)
    match = re.search(r"[（(]\s*(\d+(?:\.\d+)?)\s*分\s*[）)]", text)
    return float(match.group(1)) if match else default


def clean_item_label(label):
    text = compact_text(label)
    text = re.sub(r"^\s*[IVX]+\.\s*", "", text)
    text = re.sub(r"^\s*\d+\.\s*", "", text)
    text = re.sub(r"[（(]\s*\d+(?:\.\d+)?\s*分\s*[）)]", "", text)
    return compact_text(text)


def make_merged_lookup(ws):
    lookup = {}
    for cell_range in ws.merged_cells.ranges:
        value = ws.cell(cell_range.min_row, cell_range.min_col).value
        for row in range(cell_range.min_row, cell_range.max_row + 1):
            for col in range(cell_range.min_col, cell_range.max_col + 1):
                lookup[(row, col)] = value
    return lookup


def header_value(ws, merged_lookup, row, col):
    value = ws.cell(row, col).value
    if value not in (None, ""):
        return value
    return merged_lookup.get((row, col), "")


def section_name_from_header(value):
    text = normalize_key(value)
    if "知识部分" in text:
        return "知识部分"
    if "技能部分" in text:
        return "技能部分"
    if "态度部分" in text:
        return "态度部分"
    return ""


def unique_person_id(file_name, row):
    seq = compact_text(row.get("序号"))
    name = compact_text(row.get("姓名"))
    license_no = compact_text(row.get("驾驶员执照编号"))
    return "::".join([file_name, seq, name, license_no])


def build_column_metadata(ws):
    merged_lookup = make_merged_lookup(ws)
    metadata_cols = {}
    result_cols = {}
    detail_cols = []

    skill_item_counts = Counter()
    for col in range(1, ws.max_column + 1):
        part = section_name_from_header(header_value(ws, merged_lookup, 2, col))
        item_label = compact_text(header_value(ws, merged_lookup, 3, col))
        item_key = clean_item_label(item_label)
        if part == "技能部分" and item_key and "总分" not in item_key and "记录" not in item_key:
            skill_item_counts[item_label] += 1

    skill_item_seen = Counter()
    for col in range(1, ws.max_column + 1):
        row2 = compact_text(header_value(ws, merged_lookup, 2, col))
        row3 = compact_text(header_value(ws, merged_lookup, 3, col))
        row4 = compact_text(header_value(ws, merged_lookup, 4, col))
        part = section_name_from_header(row2)

        if col <= 10:
            label = row2 or row3 or f"列{col}"
            metadata_cols[col] = compact_text(label)
            continue

        if row3:
            key = normalize_key(row3)
            if key in {"知识部分总分", "技能部分总分", "态度部分总分", "总分数", "评估结果", "检查员评语", "评估员合格证编号", "评估员签名", "日期", "监察员审查意见", "监察员姓名"}:
                result_cols[key] = col
            elif key == "记录":
                result_cols[f"{part}记录"] = col

        if part == "知识部分":
            if normalize_key(row3) == "知识部分总分" or normalize_key(row4) != "得分":
                continue
            question_col = col - 1 if normalize_key(header_value(ws, merged_lookup, 4, col - 1)) == "题目" else None
            max_score = parse_max_score(row3, 5)
            detail_cols.append(
                {
                    "部分": part,
                    "评分项目": clean_item_label(row3),
                    "项目原名": row3,
                    "题目列": question_col,
                    "得分列": col,
                    "满分": max_score,
                    "列顺序": col,
                }
            )

        elif part == "技能部分":
            item_key = clean_item_label(row3)
            if not item_key or "总分" in item_key or item_key == "记录":
                continue
            skill_item_seen[row3] += 1
            count = max(skill_item_counts[row3], 1)
            item_max = parse_max_score(row3, np.nan)
            column_max = item_max / count if pd.notna(item_max) else np.nan
            question = item_key
            if count > 1:
                question = f"{item_key} 第{skill_item_seen[row3]}项"
            detail_cols.append(
                {
                    "部分": part,
                    "评分项目": item_key,
                    "项目原名": row3,
                    "题目列": None,
                    "固定题目": question,
                    "得分列": col,
                    "满分": column_max,
                    "列顺序": col,
                }
            )

        elif part == "态度部分":
            item_key = clean_item_label(row3)
            if not item_key or "总分" in item_key or item_key == "记录":
                continue
            max_score = parse_max_score(row3, 5)
            detail_cols.append(
                {
                    "部分": part,
                    "评分项目": item_key,
                    "项目原名": row3,
                    "题目列": None,
                    "固定题目": item_key,
                    "得分列": col,
                    "满分": max_score,
                    "列顺序": col,
                }
            )

    return metadata_cols, result_cols, detail_cols


def compute_assessment(total_score, part_scores):
    if pd.isna(total_score):
        return "未评定"
    has_low_part = any(
        pd.notna(part_scores.get(part)) and part_scores.get(part) <= max_score * 0.5
        for part, max_score in PART_MAX_SCORES.items()
    )
    if total_score >= 100 and not has_low_part:
        return "优秀"
    if total_score >= 75:
        return "通过"
    if 62 <= total_score <= 74:
        return "临界"
    return "不合格"


def normalize_result_status(value):
    text = compact_text(value)
    if not text:
        return ""
    for status in ["不合格", "优秀", "通过", "临界", "未评定"]:
        if status in text:
            return status
    return text


def compare_original_and_rule_result(original_result, rule_result):
    original_status = normalize_result_status(original_result)
    rule_status = compact_text(rule_result)
    if not original_status:
        return rule_status, "原表未填写", f"原表评估结果为空，按分数规则判定为：{rule_status}"
    if original_status == rule_status:
        return original_status, "一致", ""
    return rule_status, "不一致", f"原表评估结果：{original_status}；按分数规则：{rule_status}"


def load_evaluation_file(file_bytes, file_name):
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    metadata_cols, result_cols, detail_cols = build_column_metadata(ws)

    people_records = []
    detail_records = []

    for row_idx in range(5, ws.max_row + 1):
        row = {"来源文件": file_name, "Excel行号": row_idx}
        has_data = False
        for col, label in metadata_cols.items():
            value = ws.cell(row_idx, col).value
            row[label] = value
            if value not in (None, ""):
                has_data = True

        if not has_data or not compact_text(row.get("姓名")):
            continue

        row["人员ID"] = unique_person_id(file_name, row)
        row["知识部分总分"] = to_number(ws.cell(row_idx, result_cols.get("知识部分总分", 0)).value) if result_cols.get("知识部分总分") else np.nan
        row["技能部分总分"] = to_number(ws.cell(row_idx, result_cols.get("技能部分总分", 0)).value) if result_cols.get("技能部分总分") else np.nan
        row["态度部分总分"] = to_number(ws.cell(row_idx, result_cols.get("态度部分总分", 0)).value) if result_cols.get("态度部分总分") else np.nan
        row["技能部分记录"] = ws.cell(row_idx, result_cols.get("技能部分记录", 0)).value if result_cols.get("技能部分记录") else ""
        row["态度部分记录"] = ws.cell(row_idx, result_cols.get("态度部分记录", 0)).value if result_cols.get("态度部分记录") else ""
        row["总成绩"] = to_number(ws.cell(row_idx, result_cols.get("总分数", 0)).value) if result_cols.get("总分数") else np.nan
        row["原评估结果"] = compact_text(ws.cell(row_idx, result_cols.get("评估结果", 0)).value) if result_cols.get("评估结果") else ""
        row["检查员评语"] = compact_text(ws.cell(row_idx, result_cols.get("检查员评语", 0)).value) if result_cols.get("检查员评语") else ""
        row["评估员合格证编号"] = compact_text(ws.cell(row_idx, result_cols.get("评估员合格证编号", 0)).value) if result_cols.get("评估员合格证编号") else ""
        row["评估员签名"] = compact_text(ws.cell(row_idx, result_cols.get("评估员签名", 0)).value) if result_cols.get("评估员签名") else ""

        part_scores = {
            "知识部分": row["知识部分总分"],
            "技能部分": row["技能部分总分"],
            "态度部分": row["态度部分总分"],
        }

        for meta in detail_cols:
            score = to_number(ws.cell(row_idx, meta["得分列"]).value)
            question = meta.get("固定题目", "")
            if meta.get("题目列"):
                question = compact_text(ws.cell(row_idx, meta["题目列"]).value)
            detail_records.append(
                {
                    "人员ID": row["人员ID"],
                    "姓名": compact_text(row.get("姓名")),
                    "工作单位": compact_text(row.get("工作单位")),
                    "运行基地": compact_text(row.get("运行基地")),
                    "来源文件": file_name,
                    "部分": meta["部分"],
                    "评分项目": meta["评分项目"],
                    "题目": question,
                    "得分": score,
                    "满分": meta["满分"],
                    "得分率": score / meta["满分"] if pd.notna(score) and pd.notna(meta["满分"]) and meta["满分"] else np.nan,
                    "列顺序": meta["列顺序"],
                }
            )

        detail_for_person = [record for record in detail_records if record["人员ID"] == row["人员ID"]]
        detail_df = pd.DataFrame(detail_for_person)
        for part in PART_ORDER:
            if pd.isna(part_scores[part]) and not detail_df.empty:
                part_scores[part] = detail_df.loc[detail_df["部分"] == part, "得分"].sum(min_count=1)
                row[f"{part}总分"] = part_scores[part]

        if pd.isna(row["总成绩"]):
            numeric_part_scores = [score for score in part_scores.values() if pd.notna(score)]
            row["总成绩"] = sum(numeric_part_scores) if numeric_part_scores else np.nan

        row["安全风险提示"] = False
        row["规则评估结果"] = compute_assessment(row["总成绩"], part_scores)
        final_result, consistency, warning_text = compare_original_and_rule_result(row["原评估结果"], row["规则评估结果"])
        row["最终评估结果"] = final_result
        row["评估结果一致性"] = consistency
        row["结果差异提醒"] = warning_text
        row["人工安全风险标记"] = False
        row["总得分率"] = row["总成绩"] / TOTAL_MAX_SCORE if pd.notna(row["总成绩"]) else np.nan
        for part in PART_ORDER:
            row[f"{part}得分率"] = part_scores[part] / PART_MAX_SCORES[part] if pd.notna(part_scores[part]) else np.nan

        people_records.append(row)

    people_df = pd.DataFrame(people_records)
    detail_df = pd.DataFrame(detail_records)
    for col in PERSON_BASE_COLUMNS:
        if col not in people_df.columns:
            people_df[col] = pd.Series(dtype="object")
    for col in DETAIL_BASE_COLUMNS:
        if col not in detail_df.columns:
            detail_df[col] = pd.Series(dtype="object")
    return people_df, detail_df


def load_all_files(uploaded_files):
    people_frames = []
    detail_frames = []
    for file in uploaded_files:
        try:
            people_df, detail_df = load_evaluation_file(file.getvalue(), file.name)
            if not people_df.empty:
                people_frames.append(people_df)
                detail_frames.append(detail_df)
                st.sidebar.success(f"{file.name}: 读取 {len(people_df)} 人")
            else:
                st.sidebar.warning(f"{file.name}: 未识别到有效人员记录")
        except Exception as exc:
            st.sidebar.error(f"{file.name}: 解析失败 - {exc}")
    if not people_frames:
        return pd.DataFrame(columns=PERSON_BASE_COLUMNS), pd.DataFrame(columns=DETAIL_BASE_COLUMNS)
    return pd.concat(people_frames, ignore_index=True), pd.concat(detail_frames, ignore_index=True)


def load_default_file():
    if not DEFAULT_EXCEL_PATH.exists():
        return pd.DataFrame(columns=PERSON_BASE_COLUMNS), pd.DataFrame(columns=DETAIL_BASE_COLUMNS)
    with DEFAULT_EXCEL_PATH.open("rb") as f:
        return load_evaluation_file(f.read(), DEFAULT_EXCEL_PATH.name)


def validate_people_data(people_df):
    missing = [col for col in PERSON_REQUIRED_COLUMNS if col not in people_df.columns]
    if missing:
        st.error(
            "未能从 Excel 中识别人员基础信息列："
            + "、".join(missing)
            + "。请确认上传的是新版通航检查情况表，且表头包含“序号、姓名、工作单位”等字段。"
        )
        st.stop()
    if people_df.empty or people_df["姓名"].dropna().astype(str).str.strip().eq("").all():
        st.info("请从左侧上传新版通航检查情况 Excel 文件。")
        st.stop()


def ensure_columns(df, columns):
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            out[col] = pd.Series(dtype="object")
    return out


def result_display_label(status):
    return RESULT_LABELS.get(compact_text(status), compact_text(status) or RESULT_LABELS["未评定"])


def add_result_display_columns(people_df):
    df = people_df.copy()
    if "最终评估结果" not in df.columns:
        df["最终评估结果"] = "未评定"
    df["最终评估结果显示"] = df["最终评估结果"].map(result_display_label)
    return df


def recalculate_with_manual_safety(people_df, unsafe_names=None):
    df = people_df.copy()
    for idx, row in df.iterrows():
        part_scores = {part: row.get(f"{part}总分", np.nan) for part in PART_ORDER}
        rule_result = compute_assessment(row.get("总成绩", np.nan), part_scores)
        final_result, consistency, warning_text = compare_original_and_rule_result(row.get("原评估结果", ""), rule_result)
        df.at[idx, "规则评估结果"] = rule_result
        df.at[idx, "最终评估结果"] = final_result
        df.at[idx, "评估结果一致性"] = consistency
        df.at[idx, "结果差异提醒"] = warning_text
        df.at[idx, "安全风险提示"] = False
        df.at[idx, "人工安全风险标记"] = False
    return add_result_display_columns(df)


def filtered_data(people_df, detail_df, selected_company, selected_result):
    people = people_df.copy()
    if selected_company != "全部":
        people = people[people["工作单位"].astype(str) == selected_company]
    if selected_result != "全部":
        result_status = RESULT_LABEL_TO_STATUS.get(selected_result, selected_result)
        people = people[people["最终评估结果"].astype(str) == result_status]
    detail_df = ensure_columns(detail_df, DETAIL_BASE_COLUMNS)
    detail = detail_df[detail_df["人员ID"].isin(people["人员ID"])].copy()
    detail = ensure_columns(detail, DETAIL_BASE_COLUMNS)
    return people, detail


def company_stats(people_df):
    if people_df.empty:
        return pd.DataFrame()
    return (
        people_df.groupby("工作单位", dropna=False)
        .agg(
            人数=("姓名", "count"),
            平均分=("总成绩", "mean"),
            最高分=("总成绩", "max"),
            最低分=("总成绩", "min"),
            优秀人数=("最终评估结果", lambda s: (s == "优秀").sum()),
            通过人数=("最终评估结果", lambda s: (s == "通过").sum()),
            临界人数=("最终评估结果", lambda s: (s == "临界").sum()),
            不合格人数=("最终评估结果", lambda s: (s == "不合格").sum()),
        )
        .reset_index()
        .round({"平均分": 1, "最高分": 1, "最低分": 1})
        .sort_values(["平均分", "人数"], ascending=[False, False])
    )


def part_score_summary(people_df):
    if people_df.empty:
        return pd.DataFrame()
    rows = []
    for part in PART_ORDER:
        score_col = f"{part}总分"
        if score_col not in people_df.columns:
            continue
        values = pd.to_numeric(people_df[score_col], errors="coerce").dropna()
        rows.append(
            {
                "部分": part,
                "平均得分": values.mean() if not values.empty else np.nan,
                "满分": PART_MAX_SCORES[part],
                "平均得分率": values.mean() / PART_MAX_SCORES[part] if not values.empty else np.nan,
                "样本数": int(values.count()),
                "部分排序": PART_ORDER.index(part),
            }
        )
    return pd.DataFrame(rows).sort_values("部分排序")


def item_summary(detail_df):
    if detail_df.empty:
        return pd.DataFrame()
    summary = (
        detail_df.groupby(["部分", "评分项目"], dropna=False)
        .agg(
            平均得分=("得分", "mean"),
            平均得分率=("得分率", "mean"),
            满分=("满分", "sum"),
            样本数=("得分", "count"),
            列顺序=("列顺序", "min"),
        )
        .reset_index()
    )
    summary["平均得分率"] = summary["平均得分率"].clip(0, 1)
    summary["部分排序"] = summary["部分"].map({part: idx for idx, part in enumerate(PART_ORDER)}).fillna(len(PART_ORDER))
    return summary.sort_values(["部分排序", "列顺序"])


def fig_score_distribution(people_df):
    if people_df.empty:
        return None
    data = add_result_display_columns(people_df)
    fig = px.histogram(
        data,
        x="总成绩",
        color="最终评估结果显示",
        nbins=18,
        title="总成绩分布",
        labels={"总成绩": "总成绩", "count": "人数", "最终评估结果显示": "评估结果"},
        category_orders={"最终评估结果显示": RESULT_LABEL_ORDER},
        color_discrete_map=RESULT_COLOR_MAP,
    )
    fig.add_vline(x=100, line_dash="dash", line_color="#2563EB", annotation_text="≥100")
    fig.add_vline(x=75, line_dash="dash", line_color="#16A34A", annotation_text="≥75")
    fig.add_vline(x=62, line_dash="dash", line_color="#DC2626", annotation_text="≥62")
    return fig


def fig_result_pie(people_df):
    if people_df.empty:
        return None
    data = add_result_display_columns(people_df)
    counts = data["最终评估结果显示"].value_counts().reindex(RESULT_LABEL_ORDER, fill_value=0).reset_index()
    counts.columns = ["评估结果", "人数"]
    counts = counts[counts["人数"] > 0]
    fig = px.pie(
        counts,
        names="评估结果",
        values="人数",
        hole=0.45,
        title="评估结果构成",
        color="评估结果",
        category_orders={"评估结果": RESULT_LABEL_ORDER},
        color_discrete_map=RESULT_COLOR_MAP,
    )
    fig.update_traces(sort=False)
    return fig


def fig_company_average(people_df):
    stats = company_stats(people_df)
    if stats.empty:
        return None
    return px.bar(
        stats,
        x="工作单位",
        y="平均分",
        text="平均分",
        title="各单位平均总成绩",
        color="平均分",
        color_continuous_scale="Tealrose",
    )


def fig_person_score_bars(people_df):
    if people_df.empty:
        return None
    data = add_result_display_columns(people_df)
    data["总成绩"] = pd.to_numeric(data["总成绩"], errors="coerce")
    data = data.dropna(subset=["总成绩"]).copy()
    if data.empty:
        return None
    result_rank = {status: idx for idx, status in enumerate(RESULT_ORDER)}
    data["结果排序"] = data["最终评估结果"].map(result_rank).fillna(len(RESULT_ORDER))
    data = data.sort_values(["结果排序", "总成绩", "姓名"], ascending=[True, False, True])
    data["人员显示"] = data["姓名"].astype(str)
    duplicated = data["人员显示"].duplicated(keep=False)
    data.loc[duplicated, "人员显示"] = (
        data.loc[duplicated, "姓名"].astype(str) + "｜" + data.loc[duplicated, "工作单位"].astype(str)
    )
    person_order = data["人员显示"].tolist()
    avg_score = data["总成绩"].mean()
    fig = px.bar(
        data,
        y="人员显示",
        x="总成绩",
        orientation="h",
        color="最终评估结果显示",
        text=data["总成绩"].map(lambda value: f"{value:.1f}"),
        title="人员得分情况统计",
        labels={"总成绩": "总成绩", "人员显示": "", "最终评估结果显示": "评估结果"},
        category_orders={"最终评估结果显示": RESULT_LABEL_ORDER},
        color_discrete_map=RESULT_COLOR_MAP,
    )
    fig.add_vline(
        x=avg_score,
        line_dash="dash",
        line_color="#111827",
    )
    fig.add_annotation(
        x=avg_score,
        y=1,
        xref="x",
        yref="paper",
        text=f"平均分 {avg_score:.1f}",
        showarrow=False,
        xanchor="left",
        yanchor="bottom",
        xshift=8,
        bgcolor="rgba(255,255,255,0.94)",
        bordercolor="#111827",
        borderwidth=1,
        font=dict(color="#111827", size=14),
    )
    fig.update_traces(textposition="outside", cliponaxis=False)
    fig.update_layout(
        height=figure_height(len(data), 520, 28, 1800),
        xaxis=dict(range=[0, max(TOTAL_MAX_SCORE, data["总成绩"].max(), avg_score) * 1.08]),
        yaxis=dict(
            categoryorder="array",
            categoryarray=person_order,
            autorange="reversed",
            automargin=True,
        ),
        margin=dict(l=180, r=80, t=80, b=50),
    )
    return fig


def fig_part_average(people_df):
    summary = part_score_summary(people_df)
    if summary.empty:
        return None
    summary["平均得分"] = summary["平均得分"].round(1)
    return px.bar(
        summary,
        x="部分",
        y="平均得分",
        text="平均得分",
        title="三部分平均得分",
        category_orders={"部分": PART_ORDER},
        color="部分",
        color_discrete_sequence=["#2563EB", "#059669", "#D97706"],
    )


def fig_item_rate(detail_df):
    summary = item_summary(detail_df)
    if summary.empty:
        return None
    summary["平均得分率显示"] = (summary["平均得分率"] * 100).round(1)
    item_order = summary.sort_values(["部分排序", "列顺序"])["评分项目"].drop_duplicates().tolist()
    fig = px.bar(
        summary,
        x="评分项目",
        y="平均得分率显示",
        color="部分",
        text="平均得分率显示",
        title="各评分项目平均得分率",
        labels={"平均得分率显示": "平均得分率(%)"},
        category_orders={"部分": PART_ORDER, "评分项目": item_order},
        color_discrete_sequence=["#2563EB", "#059669", "#D97706"],
    )
    fig.update_xaxes(tickangle=-35)
    fig.add_hline(y=50, line_dash="dash", line_color="#DC2626", annotation_text="50%")
    return fig


def fig_person_radar(person_row, person_detail):
    rows = []
    detail = ensure_columns(person_detail, ["部分", "得分"]).copy()
    detail["得分"] = pd.to_numeric(detail["得分"], errors="coerce")
    radar_scale = {
        "知识部分": 60 / PART_MAX_SCORES["知识部分"],
        "技能部分": 60 / PART_MAX_SCORES["技能部分"],
        "态度部分": 60 / PART_MAX_SCORES["态度部分"],
    }
    for part in PART_ORDER:
        raw_score = pd.to_numeric(pd.Series([person_row.get(f"{part}总分", np.nan)]), errors="coerce").iloc[0]
        if pd.isna(raw_score) and not detail.empty:
            raw_score = detail.loc[detail["部分"] == part, "得分"].sum(min_count=1)
        if pd.isna(raw_score):
            continue
        rows.append(
            {
                "部分": part,
                "原始得分": raw_score,
                "原始满分": PART_MAX_SCORES[part],
                "雷达得分": raw_score * radar_scale[part],
                "雷达满分": 60,
            }
        )

    data = pd.DataFrame(rows)
    if data.empty:
        return None

    theta = data["部分"].str.replace("部分", "", regex=False).tolist()
    scores = data["雷达得分"].round(2).tolist()
    max_scores = data["雷达满分"].round(2).tolist()
    hover_text = [
        f"{part}<br>原始得分 {score:g} / {max_score:g}<br>雷达折算 {radar_score:g} / {radar_max:g}"
        for part, score, max_score, radar_score, radar_max in zip(
            data["部分"],
            data["原始得分"],
            data["原始满分"],
            data["雷达得分"],
            data["雷达满分"],
        )
    ]

    closed_theta = theta + [theta[0]]
    closed_scores = scores + [scores[0]]
    closed_max_scores = max_scores + [max_scores[0]]
    closed_hover_text = hover_text + [hover_text[0]]
    radial_max = max(max_scores) if max_scores else 10
    fig = go.Figure(
        data=[
            go.Scatterpolar(
                r=closed_max_scores,
                theta=closed_theta,
                mode="lines",
                name="折算满分",
                hovertemplate="%{theta}<br>折算满分 %{r:g}<extra></extra>",
                line=dict(color="#9CA3AF", width=2, dash="dash"),
            ),
            go.Scatterpolar(
                r=closed_scores,
                theta=closed_theta,
                fill="toself",
                name="折算得分",
                text=closed_hover_text,
                hovertemplate="%{text}<extra></extra>",
                line=dict(color="#2563EB", width=3),
                fillcolor="rgba(37, 99, 235, 0.22)",
            ),
        ]
    )
    fig.update_layout(
        title=f"{compact_text(person_row.get('姓名'))} 知识/技能/态度画像（统一折算至60分尺度）",
        polar=dict(
            radialaxis=dict(visible=True, range=[0, 60], tickvals=[0, 15, 30, 45, 60]),
            angularaxis=dict(tickfont=dict(size=14)),
        ),
        showlegend=True,
    )
    return fig


def fig_person_item_radar(person_row, person_detail):
    detail = ensure_columns(person_detail, ["部分", "评分项目", "得分", "满分", "列顺序"]).copy()
    detail["得分"] = pd.to_numeric(detail["得分"], errors="coerce")
    detail["满分"] = pd.to_numeric(detail["满分"], errors="coerce")
    detail["列顺序"] = pd.to_numeric(detail["列顺序"], errors="coerce")
    detail = detail[detail["评分项目"].notna() & detail["评分项目"].astype(str).str.strip().ne("")]
    if detail.empty:
        return None

    data = (
        detail.groupby(["部分", "评分项目"], dropna=False)
        .agg(得分=("得分", "sum"), 满分=("满分", "sum"), 列顺序=("列顺序", "min"))
        .reset_index()
    )
    data["满分"] = pd.to_numeric(data["满分"], errors="coerce")
    data = data[data["满分"].fillna(0) > 0].copy()
    if data.empty:
        return None

    data["部分排序"] = data["部分"].map({part: idx for idx, part in enumerate(PART_ORDER)})
    data = data.sort_values(["部分排序", "列顺序", "评分项目"])
    data["雷达轴"] = data["部分"].str.replace("部分", "", regex=False) + "｜" + data["评分项目"].astype(str)
    data["原始得分"] = data["得分"].fillna(0)
    data["折算得分"] = (data["原始得分"] / data["满分"] * 20).clip(lower=0, upper=20)
    data["折算满分"] = 20

    theta = data["雷达轴"].tolist()
    scores = data["折算得分"].round(2).tolist()
    max_scores = data["折算满分"].round(2).tolist()
    hover_text = [
        f"{part}<br>{item}<br>原始得分 {score:g} / {max_score:g}<br>统一20分折算 {scaled:g} / 20"
        for part, item, score, max_score, scaled in zip(
            data["部分"],
            data["评分项目"],
            data["原始得分"],
            data["满分"],
            data["折算得分"],
        )
    ]

    closed_theta = theta + [theta[0]]
    closed_scores = scores + [scores[0]]
    closed_max_scores = max_scores + [max_scores[0]]
    closed_hover_text = hover_text + [hover_text[0]]
    fig = go.Figure(
        data=[
            go.Scatterpolar(
                r=closed_max_scores,
                theta=closed_theta,
                mode="lines",
                name="统一满分",
                hovertemplate="%{theta}<br>统一满分 %{r:g}<extra></extra>",
                line=dict(color="#9CA3AF", width=2, dash="dash"),
            ),
            go.Scatterpolar(
                r=closed_scores,
                theta=closed_theta,
                fill="toself",
                name="折算得分",
                text=closed_hover_text,
                hovertemplate="%{text}<extra></extra>",
                line=dict(color="#059669", width=3),
                fillcolor="rgba(5, 150, 105, 0.20)",
            ),
        ]
    )
    fig.update_layout(
        title=f"{compact_text(person_row.get('姓名'))} 评分项目雷达图（统一20分尺度）",
        polar=dict(
            radialaxis=dict(visible=True, range=[0, 20], tickvals=[0, 5, 10, 15, 20]),
            angularaxis=dict(tickfont=dict(size=11)),
        ),
        showlegend=True,
    )
    return fig


def fig_person_items(person_detail):
    if person_detail.empty:
        return None
    data = person_detail.copy()
    data["得分"] = pd.to_numeric(data["得分"], errors="coerce")
    data["满分"] = pd.to_numeric(data["满分"], errors="coerce")
    data["项目显示"] = data["评分项目"]
    duplicated = data["项目显示"].duplicated(keep=False)
    data["项目序号"] = data.groupby("评分项目", dropna=False).cumcount() + 1
    data.loc[duplicated, "项目显示"] = (
        data.loc[duplicated, "评分项目"].astype(str)
        + " 第"
        + data.loc[duplicated, "项目序号"].astype(str)
        + "项"
    )
    data["得分文本"] = data.apply(
        lambda row: f"{row['得分']:g}/{row['满分']:g}" if pd.notna(row["得分"]) and pd.notna(row["满分"]) else "",
        axis=1,
    )
    fig = px.bar(
        data.sort_values("列顺序"),
        x="项目显示",
        y="得分",
        color="部分",
        text="得分文本",
        title="个人评分项目得分",
        labels={"得分": "得分"},
        color_discrete_sequence=["#2563EB", "#059669", "#D97706"],
    )
    fig.update_xaxes(tickangle=-35)
    return fig


def display_people_table(people_df):
    people_df = add_result_display_columns(people_df)
    cols = [
        "序号",
        "姓名",
        "驾驶员执照编号",
        "工作单位",
        "运行基地",
        "知识部分总分",
        "技能部分总分",
        "态度部分总分",
        "总成绩",
        "最终评估结果显示",
        "原评估结果",
        "规则评估结果",
        "评估结果一致性",
        "结果差异提醒",
        "检查员评语",
    ]
    cols = [col for col in cols if col in people_df.columns]
    return people_df[cols].sort_values("总成绩", ascending=False)


st.title("通航驾驶员能力评估分析平台")
st.caption("按新版 125 分调研工作单解析：知识部分 40 分、技能部分 60 分、态度部分 25 分。")

with st.sidebar:
    st.markdown("### 数据文件")
    uploaded_files = st.file_uploader("上传新版评估 Excel", type=["xlsx", "xls"], accept_multiple_files=True)
    use_default = st.checkbox("未上传时使用当前目录示例数据", value=DEFAULT_EXCEL_PATH.exists())
    st.markdown("### 评估规则")
    st.markdown(
        """
        - 总分：125 分
        - 优秀：总成绩 ≥ 100，且知识/技能/态度无单项 ≤ 50%
        - 通过：总成绩 ≥ 75
        - 临界：总成绩 ≥ 62
        - 不合格：总成绩 < 62
        - 系统会将分数规则结果与原表“评估结果”进行对比，不一致时提醒关注
        """
    )

if uploaded_files:
    people_data, detail_data = load_all_files(uploaded_files)
elif use_default and DEFAULT_EXCEL_PATH.exists():
    try:
        people_data, detail_data = load_default_file()
        st.sidebar.success(f"已加载示例数据：{DEFAULT_EXCEL_PATH.name}")
    except Exception as exc:
        st.sidebar.error(f"示例数据解析失败：{exc}")
        people_data, detail_data = pd.DataFrame(columns=PERSON_BASE_COLUMNS), pd.DataFrame(columns=DETAIL_BASE_COLUMNS)
else:
    people_data, detail_data = pd.DataFrame(columns=PERSON_BASE_COLUMNS), pd.DataFrame(columns=DETAIL_BASE_COLUMNS)

validate_people_data(people_data)

people_data = recalculate_with_manual_safety(people_data)
detail_data = detail_data.merge(
    people_data[["人员ID", "最终评估结果"]],
    on="人员ID",
    how="left",
)

with st.sidebar:
    companies = ["全部"] + sorted(people_data["工作单位"].dropna().astype(str).unique().tolist())
    selected_company = st.selectbox("单位筛选", companies)
    existing_results = set(people_data["最终评估结果"].dropna().astype(str))
    result_options = ["全部"] + [RESULT_LABELS[item] for item in RESULT_ORDER if item in existing_results]
    selected_result = st.selectbox("评估结果筛选", result_options)

filtered_people, filtered_detail = filtered_data(people_data, detail_data, selected_company, selected_result)

st.success(
    f"成功加载 {len(people_data)} 人，识别 {len(detail_data)} 条评分明细；当前筛选后 {len(filtered_people)} 人。"
)

mismatch_people = filtered_people[filtered_people["评估结果一致性"].astype(str).isin(["不一致", "原表未填写"])].copy()
if not mismatch_people.empty:
    st.warning(f"当前筛选范围内有 {len(mismatch_people)} 人的原表评估结果与分数规则结果不一致，请重点核对。")
    st.dataframe(
        mismatch_people[["姓名", "工作单位", "总成绩", "原评估结果", "规则评估结果", "结果差异提醒"]],
        use_container_width=True,
        hide_index=True,
    )

tab_overview, tab_parts, tab_person, tab_export = st.tabs(["整体情况", "三部分分析", "个人详情", "数据导出"])

with tab_overview:
    valid_scores = filtered_people["总成绩"].dropna()
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("人数", len(filtered_people))
    col2.metric("平均分", f"{valid_scores.mean():.1f}" if not valid_scores.empty else "-")
    col3.metric("最高分", f"{valid_scores.max():.1f}" if not valid_scores.empty else "-")
    col4.metric("最低分", f"{valid_scores.min():.1f}" if not valid_scores.empty else "-")
    col5.metric("优秀率", f"{(filtered_people['最终评估结果'].eq('优秀').mean() * 100):.1f}%" if len(filtered_people) else "-")

    chart_col1, chart_col2 = st.columns(2)
    with chart_col1:
        fig = fig_score_distribution(filtered_people)
        if fig:
            st.plotly_chart(fig, use_container_width=True)
    with chart_col2:
        fig = fig_result_pie(filtered_people)
        if fig:
            st.plotly_chart(fig, use_container_width=True)

    fig = fig_company_average(filtered_people)
    if fig:
        st.plotly_chart(fig, use_container_width=True)

    fig = fig_person_score_bars(filtered_people)
    if fig:
        st.plotly_chart(fig, use_container_width=True)

    st.markdown("#### 各单位统计")
    st.dataframe(company_stats(filtered_people), use_container_width=True, hide_index=True)

with tab_parts:
    chart_col1, chart_col2 = st.columns([1, 2])
    with chart_col1:
        fig = fig_part_average(filtered_people)
        if fig:
            st.plotly_chart(fig, use_container_width=True)
    with chart_col2:
        fig = fig_item_rate(filtered_detail)
        if fig:
            st.plotly_chart(fig, use_container_width=True)

    st.markdown("#### 评分项目明细汇总")
    summary = item_summary(filtered_detail)
    if not summary.empty:
        shown = summary[["部分", "评分项目", "平均得分", "平均得分率", "样本数"]].copy()
        shown["平均得分"] = shown["平均得分"].round(2)
        shown["平均得分率"] = (shown["平均得分率"] * 100).round(1).astype(str) + "%"
        st.dataframe(shown, use_container_width=True, hide_index=True)

    with st.expander("查看全部题目及分数明细"):
        detail_cols = ["姓名", "工作单位", "部分", "评分项目", "题目", "得分", "满分", "得分率"]
        detail_show = ensure_columns(filtered_detail, detail_cols)[detail_cols].copy()
        if detail_show.empty:
            st.info("当前筛选条件下暂无题目明细。")
        else:
            detail_show["得分率"] = (pd.to_numeric(detail_show["得分率"], errors="coerce") * 100).round(1).astype(str) + "%"
            st.dataframe(detail_show.sort_values(["姓名", "部分", "评分项目"]), use_container_width=True, hide_index=True)

with tab_person:
    st.markdown("#### 个人画像筛选")
    if filtered_people.empty:
        st.info("当前筛选条件下没有人员。")
    else:
        name_options = filtered_people.sort_values("总成绩", ascending=False)["姓名"].dropna().astype(str).tolist()
        selected_name = st.selectbox("选择人员", name_options)
        person_rows = filtered_people[filtered_people["姓名"].astype(str) == selected_name]
        person = person_rows.iloc[0]
        person_detail = filtered_detail[filtered_detail["人员ID"] == person["人员ID"]].copy()

        metric_col1, metric_col2, metric_col3, metric_col4, metric_col5 = st.columns(5)
        metric_col1.metric("总成绩", f"{person.get('总成绩', np.nan):.1f}" if pd.notna(person.get("总成绩")) else "-")
        metric_col2.metric("评估结果", person.get("最终评估结果显示", result_display_label(person.get("最终评估结果", "-"))))
        metric_col3.metric("知识/40", f"{person.get('知识部分总分', np.nan):.1f}" if pd.notna(person.get("知识部分总分")) else "-")
        metric_col4.metric("技能/60", f"{person.get('技能部分总分', np.nan):.1f}" if pd.notna(person.get("技能部分总分")) else "-")
        metric_col5.metric("态度/25", f"{person.get('态度部分总分', np.nan):.1f}" if pd.notna(person.get("态度部分总分")) else "-")

        chart_col1, chart_col2 = st.columns(2)
        with chart_col1:
            fig = fig_person_radar(person, person_detail)
            if fig:
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("该人员暂无可绘制的评分项目画像。")
        with chart_col2:
            fig = fig_person_item_radar(person, person_detail)
            if fig:
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("该人员暂无可绘制的评分项目雷达图。")

        fig = fig_person_items(person_detail)
        if fig:
            st.plotly_chart(fig, use_container_width=True)

        st.markdown("#### 个人题目及分数")
        person_detail_cols = ["部分", "评分项目", "题目", "得分", "满分", "得分率"]
        detail_show = ensure_columns(person_detail, person_detail_cols)[person_detail_cols].copy()
        if detail_show.empty:
            st.info("该人员暂无题目明细。")
        else:
            detail_show["得分率"] = (pd.to_numeric(detail_show["得分率"], errors="coerce") * 100).round(1).astype(str) + "%"
            st.dataframe(detail_show.sort_values(["部分", "评分项目"]), use_container_width=True, hide_index=True)

        st.markdown("#### 检查员评语")
        st.write(person.get("检查员评语") or "无")

with tab_export:
    st.markdown("### 数据预览")
    st.dataframe(display_people_table(filtered_people), use_container_width=True, hide_index=True)

    people_csv = display_people_table(people_data).to_csv(index=False, encoding="utf-8-sig")
    detail_csv = detail_data.to_csv(index=False, encoding="utf-8-sig")
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            "下载人员总表 CSV",
            data=people_csv,
            file_name="通航驾驶员能力评估_人员总表.csv",
            mime="text/csv",
        )
    with col2:
        st.download_button(
            "下载题目明细 CSV",
            data=detail_csv,
            file_name="通航驾驶员能力评估_题目明细.csv",
            mime="text/csv",
        )
